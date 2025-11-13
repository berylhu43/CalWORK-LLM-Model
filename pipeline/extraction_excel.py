import os
import json
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, '..', 'data', 'dashboard_raw_data.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, '..', 'data', 'jsonl')

os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_table_streaming(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    for sheet in sheet_names:
        ws = wb[sheet]
        print(f"Processing sheet: {sheet}")

        output_file = os.path.join(OUTPUT_DIR, f"unstructured_{sheet}.jsonl")
        with open(output_file, "w", encoding="utf-8") as f:

            header = None
            row_number = 0

            for idx, row in enumerate(ws.iter_rows(values_only=True, max_col=10)):
                row_number += 1
                if row_number == 1:
                    # Skip the first row unconditionally (description or blank)
                    continue
                if row_number == 2:
                    # Take the second row as header
                    header = list(row)
                    continue

                # 跳过空行
                if all(v is None for v in row):
                    continue

                # 构造记录
                record_text = "; ".join(
                    f"{col}: {val}"
                    for col, val in zip(header, row)
                    if val is not None
                )

                record = {
                    "text": record_text,
                    "metadata": {
                        "sheet": sheet,
                        "row_index": idx
                    }
                }

                f.write(json.dumps(record, ensure_ascii=False) + "\n")

        print(f"Done: {output_file}")


if __name__ == "__main__":
    load_table_streaming(XLSX_PATH)

